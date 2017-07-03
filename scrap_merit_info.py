import xlsxwriter
from openpyxl import load_workbook
import os
import re
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
import logging
import logging.handlers
from datetime import datetime
import argparse


# logging
LOG_FILENAME = "scrap_merit_info.out"

my_logger = logging.getLogger('ScrapMeritInfoLogger')
my_logger.setLevel(logging.DEBUG)

# Add the log message handler to the logger
handler = logging.handlers.RotatingFileHandler(
              LOG_FILENAME, maxBytes=10*1024*1024, backupCount=5)

my_logger.addHandler(handler)

my_logger.info('\n\n-----------------------------------------------------')
my_logger.info('-----------------------------------------------------')
my_logger.info('Time of Execution: {}'.format(datetime.now()))


URL = "http://www.jacpcldce.ac.in/search/WPQuery_13.asp"

# FILENAMES = []

bad_chars = r'\xa0\t\n\\'
rgx = re.compile('[%s]' % bad_chars)


def main(filename):

    try:
        my_logger.info("open(in_file): {}".format(filename))
        in_file = load_workbook(filename)
        first_sheet = in_file.get_sheet_names()[0]
        worksheet = in_file.get_sheet_by_name(first_sheet)

        my_logger.info("open(out_file): {}".format("final_" + filename))
        out_file = xlsxwriter.Workbook("final_" + filename)
        sheet = out_file.add_worksheet()

        row = 1
        col = 0

        sheet.write('A1', 'Merit No')
        sheet.write('B1', 'Gujcet No')
        sheet.write('C1', 'Name')
        sheet.write('D1', 'Board Name')
        sheet.write('E1', 'Board PCM Toatl')
        sheet.write('F1', 'Gujcet PCM Total')
        sheet.write('G1', 'PCM Board Percntile (A)')
        sheet.write('H1', 'PCM Board Percntile (B)')
        sheet.write('I1', 'Merit Mark (A*0.6 + B*0.4)')
        sheet.write('J1', 'Remarks')
        sheet.write('K1', 'Alloted Institute')
        sheet.write('L1', 'Course alloted')
        sheet.write('M1', 'Alloted Category')
        sheet.write('N1', 'Sub Category')


        driver = webdriver.Chrome()
        driver.get(URL)

        for gids in worksheet.iter_rows():

            gid = gids[0].value

            my_logger.info("Getting merit info for gid: {}".format(gid))

            elm = driver.find_element_by_name('txtGcetNo')
            elm.clear()
            elm.send_keys(str(gid) + Keys.RETURN)

            soup = BeautifulSoup(driver.page_source, 'html.parser')

            tables = soup.find_all('table', limit=5)
            if len(tables) < 5:
                driver.back()
                continue

            rows = tables[-1].find_all('tr')
            data = [[rgx.sub('', td.text) for td in tr.findAll("td")] for tr in rows]

            for i in range(14):
                sheet.write(row, col, data[i][1])
                col += 1

            row += 1
            col = 0

            driver.back()
            sleep(0.05)

    except KeyboardInterrupt:
        sys.exit(0)

    finally:
        if in_file:
            in_file.close()
        if out_file:
            out_file.close()
        if driver:
            driver.close()


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Scrap Merit Info from ACPC')
    parser.add_argument('--filenames', nargs='*', required=True,
            help='filenames for the input')

    args = parser.parse_args()

    for f in args.filenames:
        main(f)
